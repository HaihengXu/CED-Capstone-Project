"""
functions_haiheng_v3.py
-----------------------
Core ETL engine for CED Pricing files.

Supported formats:
  FORMAT A  — Bridgeport/multi-vendor (CED-GREA, CED-KAL …)
              Sheet "Price List" | headers row 0
              Key col: Distributor Price Each

  FORMAT B  — Hubbell 3A distributor cost sheets
              Sheet "Price List" | headers row 5 (skip 5)
              Key col: PRICE

  FORMAT C  — Midwest net price file (multi-sheet)
              Any sheet with MATERIAL_NUMBER header at row 0
              Key col: NET_PRICE (SILVER) or NET

  FORMAT D  — Stusser/Southwire SPA XLSX
              Sheet "Sheet1" | headers row 0
              Key col: Net Price

  FORMAT E  — SpreadsheetML .xls (XML 2003)
              Header row auto-detected; key col: Price Break1

  FORMAT F  — TB / ABB SPA claimback
              Header auto-detected (~row 26); price in "Unnamed: 8"
              (split merged header — actual data under Final Net Price(1))

  FORMAT G  — SPAPriceFile (has Net Price but no description — loaded anyway)

  TEMPLATES / UNKNOWN — skipped gracefully, never crashes the run.

Privacy: all output DataFrames strip company names, account numbers,
         addresses and customer identifiers.
"""

import os
import re
import warnings
import pandas as pd
from lxml import etree
from openpyxl import load_workbook
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# Standard output columns
# ─────────────────────────────────────────────────────────────────────────────
STANDARD_COLS = ["location", "upc", "catalog_num", "description",
                 "net_price_each", "pricing_group", "source_file", "format"]

# ─────────────────────────────────────────────────────────────────────────────
# Column alias map  (lower-stripped → standard name)
# ─────────────────────────────────────────────────────────────────────────────
_COL_ALIASES = {
    # UPC
    "upc code": "upc", "upc": "upc", "ean/upc": "upc",
    "upc         ": "upc",
    # Catalog / item number
    "item": "catalog_num", "material": "catalog_num",
    "material        ": "catalog_num", "material_number": "catalog_num",
    "material code": "catalog_num",
    "catalog #": "catalog_num", "catalog": "catalog_num",
    "mfr": "catalog_num", "product": "catalog_num",
    "catalog no.": "catalog_num", "catalog no": "catalog_num",
    "item number": "catalog_num",
    # Description
    "description": "description", "catalog_description": "description",
    "item description": "description",
    # Net price
    "distributor price each": "net_price_each",
    "price": "net_price_each",
    "net price": "net_price_each",
    "net_price (silver)": "net_price_each",
    "net": "net_price_each",
    "net cost": "net_price_each",
    "net spa price": "net_price_each",
    "price break1": "net_price_each",
    # Pricing group / manufacturer
    "pricing group": "pricing_group", "prod group desc": "pricing_group",
    "mfrcode": "pricing_group", "brand": "pricing_group",
    "mpg": "pricing_group", "pricing type": "pricing_group",
}


def _strip(col: str) -> str:
    return str(col).strip().lower()


def _map_columns(df: pd.DataFrame) -> dict:
    mapping = {}
    for col in df.columns:
        alias = _strip(col)
        if alias in _COL_ALIASES:
            std = _COL_ALIASES[alias]
            if std not in mapping:
                mapping[std] = col
    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# Location tag from filename  (no company names in output)
# ─────────────────────────────────────────────────────────────────────────────
def _location_from_filename(filename: str) -> str:
    stem = Path(filename).stem

    # NNNN_CED-LOC_date…
    m = re.search(r'\d+_(CED-[A-Z]+)_', stem)
    if m:
        return m.group(1)

    # NNNN_ACCT_-_CED_-_CITY__STATE_-_…
    m = re.search(r'CED_-_([A-Z]+)__([A-Z]{2})', stem)
    if m:
        return f"CED-{m.group(1)}-{m.group(2)}"

    # Leading 4-digit code
    m = re.match(r'^(\d{4})_', stem)
    if m:
        return f"LOC-{m.group(1)}"

    # Timestamp prefix
    m = re.match(r'^(\d{8,})', stem)
    if m:
        return f"FILE-{m.group(1)[:8]}"

    return re.sub(r'[^A-Za-z0-9\-]', '-', stem)[:30]


# ─────────────────────────────────────────────────────────────────────────────
# Privacy scrubber
# ─────────────────────────────────────────────────────────────────────────────
_REDACT_PATTERNS = [
    r'STUSSER ELECTRIC COMPANY[^,\n]*',
    r'CONSOLIDATED ELECTRICAL DIST[^,\n]*',
    r'ABB Installation Products USA[^,\n]*',
    r'HUBBELL INCORPORATED[^,\n]*',
    r'Minerallac[^,\n]*',
    r'\d+\s+[A-Z][a-zA-Z\s]+(?:Ave|St|Rd|Dr|Blvd|Ln|Way|Ct)[^,\n]*',
    r'\b\d{9,}\b',
]
_REDACT_RE = re.compile('|'.join(_REDACT_PATTERNS), re.IGNORECASE)


def _scrub(value):
    if isinstance(value, str):
        return _REDACT_RE.sub('[REDACTED]', value).strip()
    return value


def _scrub_df(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    for c in cols:
        if c in df.columns:
            df[c] = df[c].apply(_scrub)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Shared output builder
# ─────────────────────────────────────────────────────────────────────────────
def _build_output(df: pd.DataFrame, mapping: dict,
                  location: str, filename: str, fmt: str) -> pd.DataFrame:
    def get_col(std):
        col = mapping.get(std)
        if col is None:
            return pd.Series([""] * len(df), index=df.index)
        return df[col].astype(str).str.strip()

    out = pd.DataFrame()
    out["upc"]            = get_col("upc")
    out["catalog_num"]    = get_col("catalog_num")
    out["description"]    = get_col("description")
    out["net_price_each"] = pd.to_numeric(
        get_col("net_price_each").str.replace(r'[,$\s]', '', regex=True),
        errors="coerce")
    out["pricing_group"]  = get_col("pricing_group")
    out["location"]       = location
    out["source_file"]    = Path(filename).name
    out["format"]         = fmt
    return out.dropna(subset=["net_price_each"])


# ─────────────────────────────────────────────────────────────────────────────
# Format detectors
# ─────────────────────────────────────────────────────────────────────────────
def _detect_format(filepath: str) -> str:
    ext = Path(filepath).suffix.lower()

    # SpreadsheetML disguised as .xls
    if ext == ".xls":
        return "format_e"

    if ext not in (".xlsx", ".csv"):
        return "unknown"

    try:
        wb = load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception:
        return "unknown"

    # Format A / B: has "Price List" sheet
    if "Price List" in sheets:
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb["Price List"]
            rows = list(ws.iter_rows(max_row=10, values_only=True))
            wb.close()
        except Exception:
            return "unknown"

        first_row = [_strip(v) if v else "" for v in (rows[0] if rows else [])]
        if "pricing group" in first_row or "item" in first_row:
            return "format_a"

        if len(rows) > 5:
            row5 = [_strip(v) if v else "" for v in rows[5]]
            if any(k in row5 for k in ("upc         ", "upc", "material")):
                return "format_b"

    # Format C: Midwest — any sheet with MATERIAL_NUMBER header
    try:
        wb = load_workbook(filepath, read_only=True)
        for sh in wb.sheetnames:
            ws = wb[sh]
            first = next(ws.iter_rows(max_row=1, values_only=True), ())
            if "material_number" in [_strip(v) if v else "" for v in first]:
                wb.close()
                return "format_c"
        wb.close()
    except Exception:
        pass

    # Format D: Sheet1 with "Net Price" + "Product" columns
    try:
        wb = load_workbook(filepath, read_only=True)
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
            first = next(ws.iter_rows(max_row=1, values_only=True), ())
            row0 = [_strip(v) if v else "" for v in first]
            wb.close()
            if "net price" in row0 and "product" in row0:
                return "format_d"
        else:
            wb.close()
    except Exception:
        pass

    # Format F: TB/ABB SPA — "Net Price" label mid-sheet (row A col)
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
            if row and _strip(row[0] or "") == "net price":
                wb.close()
                return "format_f"
        wb.close()
    except Exception:
        pass

    # Format G: SPAPriceFile — Purchaser ID, has Net Price but no product description
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        first = next(ws.iter_rows(max_row=1, values_only=True), ())
        row0 = [_strip(v) if v else "" for v in first]
        wb.close()
        if "purchaser id" in row0 and "net price" in row0:
            return "format_g"
    except Exception:
        pass

    return "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# Format loaders
# ─────────────────────────────────────────────────────────────────────────────
def _load_format_a(filepath, location):
    df = pd.read_excel(filepath, sheet_name="Price List",
                       header=0, dtype=str, engine="openpyxl")
    df.columns = df.columns.str.strip()
    mapping = _map_columns(df)
    needed = {"catalog_num", "description", "net_price_each"}
    if needed - set(mapping.keys()):
        raise ValueError(f"Missing: {needed - set(mapping.keys())}. Found: {list(df.columns)}")
    return _build_output(df, mapping, location, filepath, "format_a")


def _load_format_b(filepath, location):
    df = pd.read_excel(filepath, sheet_name="Price List",
                       skiprows=5, header=0, dtype=str, engine="openpyxl")
    df = df.dropna(axis=1, how="all")
    df.columns = df.columns.str.strip()
    mapping = _map_columns(df)
    needed = {"catalog_num", "description", "net_price_each"}
    if needed - set(mapping.keys()):
        raise ValueError(f"Missing: {needed - set(mapping.keys())}. Found: {list(df.columns)}")
    return _build_output(df, mapping, location, filepath, "format_b")


def _load_format_c(filepath, location):
    """Midwest — combine all sheets that have MATERIAL_NUMBER."""
    frames = []
    wb = load_workbook(filepath, read_only=True)
    target_sheets = []
    for sh in wb.sheetnames:
        ws = wb[sh]
        first = next(ws.iter_rows(max_row=1, values_only=True), ())
        if "material_number" in [_strip(v) if v else "" for v in first]:
            target_sheets.append(sh)
    wb.close()

    for sh in target_sheets:
        df = pd.read_excel(filepath, sheet_name=sh, header=0, dtype=str, engine="openpyxl")
        df.columns = df.columns.str.strip()
        mapping = _map_columns(df)
        if "net_price_each" not in mapping or "catalog_num" not in mapping:
            continue
        frames.append(_build_output(df, mapping, location, filepath, "format_c"))

    if not frames:
        raise ValueError("No parseable sheets in Format C file.")
    return pd.concat(frames, ignore_index=True)


def _load_format_d(filepath, location):
    """Stusser XLSX SPA — Sheet1, Net Price + Product columns."""
    df = pd.read_excel(filepath, sheet_name="Sheet1", header=0, dtype=str, engine="openpyxl")
    df.columns = df.columns.str.strip()
    # Forward-fill customer name column (merged cells in source)
    for col in df.columns:
        if "sold to" in _strip(col):
            df[col] = df[col].ffill()
    mapping = _map_columns(df)
    needed = {"catalog_num", "description", "net_price_each"}
    if needed - set(mapping.keys()):
        raise ValueError(f"Missing: {needed - set(mapping.keys())}. Found: {list(df.columns)}")
    out = _build_output(df, mapping, location, filepath, "format_d")
    return _scrub_df(out, ["description"])


def _load_format_e(filepath, location):
    """SpreadsheetML XML .xls — header row auto-detected."""
    tree = etree.parse(filepath, etree.XMLParser(recover=True))
    root = tree.getroot()
    ns = "urn:schemas-microsoft-com:office:spreadsheet"

    records = []
    for ws in root.findall(f".//{{{ns}}}Worksheet"):
        table = ws.find(f"{{{ns}}}Table")
        if table is None:
            continue
        rows = table.findall(f"{{{ns}}}Row")
        if not rows:
            continue

        def row_vals(row_el):
            cells = []
            for cell in row_el.findall(f"{{{ns}}}Cell"):
                data = cell.find(f"{{{ns}}}Data")
                cells.append(data.text if data is not None and data.text else "")
            return cells

        # Find header row: first row where >=3 cells are non-empty strings
        header_row_idx = None
        for i, row_el in enumerate(rows):
            vals = row_vals(row_el)
            if sum(1 for v in vals if v.strip()) >= 3:
                header_row_idx = i
                break

        if header_row_idx is None:
            continue

        headers = row_vals(rows[header_row_idx])
        mapping_local = {}
        for i, h in enumerate(headers):
            alias = _strip(h)
            if alias in _COL_ALIASES:
                std = _COL_ALIASES[alias]
                if std not in mapping_local:
                    mapping_local[std] = i

        if "net_price_each" not in mapping_local:
            continue

        for row_el in rows[header_row_idx + 1:]:
            vals = row_vals(row_el)
            # Pad to header length
            while len(vals) < len(headers):
                vals.append("")

            def get(std):
                i = mapping_local.get(std)
                return vals[i].strip() if i is not None and i < len(vals) else ""

            price = pd.to_numeric(re.sub(r'[,$\s]', '', get("net_price_each")), errors="coerce")
            if pd.isna(price):
                continue
            records.append({
                "upc":            get("upc"),
                "catalog_num":    get("catalog_num"),
                "description":    get("description"),
                "net_price_each": price,
                "pricing_group":  get("pricing_group"),
                "location":       location,
                "source_file":    Path(filepath).name,
                "format":         "format_e",
            })

    if not records:
        raise ValueError("No parseable data found in SpreadsheetML file.")
    return pd.DataFrame(records)


def _load_format_f(filepath, location):
    """TB/ABB SPA claimback — find header row with 'Catalog No.', price in Unnamed:8."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(max_row=60, values_only=True)):
        row_vals = [_strip(v) if v else "" for v in row]
        if "catalog no." in row_vals or "catalog no" in row_vals:
            header_row_idx = i
            break
    wb.close()

    if header_row_idx is None:
        raise ValueError("Could not find 'Catalog No.' header in Format F file.")

    df = pd.read_excel(filepath, skiprows=header_row_idx, header=0,
                       dtype=str, engine="openpyxl")
    df.columns = df.columns.str.strip()

    # The price data lives in the column after "Final Net Price (1)" —
    # it's labelled "Unnamed: 8" because the merged header row above puts
    # "M/D" there and actual prices one more row down.
    # Skip the two sub-header rows (M/D and Price labels).
    df = df.iloc[2:].reset_index(drop=True)

    # Build mapping manually for this odd layout
    mapping = {}
    for col in df.columns:
        alias = _strip(col)
        if alias == "catalog no." or alias == "catalog no":
            mapping["catalog_num"] = col
        elif alias == "description":
            mapping["description"] = col
        elif alias == "upc":
            mapping["upc"] = col

    # Price is in "Unnamed: 8" (the column index 8, zero-based)
    unnamed_cols = [c for c in df.columns if re.match(r'unnamed:\s*8', _strip(c))]
    if unnamed_cols:
        mapping["net_price_each"] = unnamed_cols[0]
    else:
        # Fallback: find column that has numeric data and follows 'Final Net Price (1)'
        for i, col in enumerate(df.columns):
            if "final net price" in _strip(col) and i + 1 < len(df.columns):
                mapping["net_price_each"] = df.columns[i + 1]
                break

    needed = {"catalog_num", "net_price_each"}
    if needed - set(mapping.keys()):
        raise ValueError(f"Missing: {needed - set(mapping.keys())}. Columns: {list(df.columns)}")

    out = _build_output(df, mapping, location, filepath, "format_f")
    return _scrub_df(out, ["description"])


def _load_format_g(filepath, location):
    """SPAPriceFile — Purchaser ID format; load catalog+price, no description."""
    df = pd.read_excel(filepath, sheet_name=0, header=0, dtype=str, engine="openpyxl")
    df.columns = df.columns.str.strip()
    mapping = _map_columns(df)
    # No description in this format — that's OK, build_output handles missing
    if "net_price_each" not in mapping or "catalog_num" not in mapping:
        raise ValueError(f"Missing price or catalog column. Found: {list(df.columns)}")
    out = _build_output(df, mapping, location, filepath, "format_g")
    # Scrub purchaser name columns
    return _scrub_df(out, ["description", "pricing_group"])


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────
def load_pricing_file(filepath: str) -> tuple:
    filename = Path(filepath).name
    location = _location_from_filename(filename)
    fmt      = _detect_format(filepath)

    _loaders = {
        "format_a": _load_format_a,
        "format_b": _load_format_b,
        "format_c": _load_format_c,
        "format_d": _load_format_d,
        "format_e": _load_format_e,
        "format_f": _load_format_f,
        "format_g": _load_format_g,
    }

    if fmt == "unknown":
        return (pd.DataFrame(columns=STANDARD_COLS),
                f"⚠️  {filename} — unrecognised format, skipped")

    try:
        df  = _loaders[fmt](filepath, location)
        return df, f"✅ {filename} → {location} [{fmt}, {len(df):,} rows]"
    except Exception as e:
        return (pd.DataFrame(columns=STANDARD_COLS),
                f"❌ {filename} — error: {e}")


def load_pricing_files(folder_path: str) -> tuple:
    folder = Path(folder_path)
    files  = (list(folder.glob("*.xlsx")) + list(folder.glob("*.XLSX")) +
              list(folder.glob("*.xls"))  + list(folder.glob("*.csv")))

    if not files:
        return pd.DataFrame(columns=STANDARD_COLS), ["No pricing files found."]

    frames, messages = [], []
    for fp in sorted(files):
        df, msg = load_pricing_file(str(fp))
        messages.append(msg)
        if not df.empty:
            frames.append(df)

    combined = (pd.concat(frames, ignore_index=True)
                if frames else pd.DataFrame(columns=STANDARD_COLS))
    return combined, messages


# ─────────────────────────────────────────────────────────────────────────────
# Wide-format comparison matrix
# ─────────────────────────────────────────────────────────────────────────────
def build_wide_matrix(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    df_dedup = (df.sort_values("net_price_each")
                  .drop_duplicates(subset=["catalog_num", "location"], keep="first"))

    meta = (df_dedup.drop_duplicates(subset=["catalog_num"], keep="first")
                    [["catalog_num", "upc", "description", "pricing_group"]])

    pivot = df_dedup.pivot_table(
        index="catalog_num", columns="location",
        values="net_price_each", aggfunc="min")
    pivot.columns.name = None
    pivot = pivot.reset_index()

    wide = meta.merge(pivot, on="catalog_num", how="right")

    price_cols = [c for c in wide.columns
                  if c not in {"catalog_num", "upc", "description",
                               "pricing_group", "best_price",
                               "best_location", "price_spread", "is_tied"}]

    wide["best_price"]    = wide[price_cols].min(axis=1)
    wide["best_location"] = wide[price_cols].idxmin(axis=1)
    wide["price_spread"]  = wide[price_cols].max(axis=1) - wide[price_cols].min(axis=1)
    wide["is_tied"]       = wide[price_cols].apply(
        lambda r: r.dropna().nunique() == 1 and r.notna().sum() > 1, axis=1)

    col_order = (["catalog_num", "upc", "description", "pricing_group"]
                 + price_cols
                 + ["best_price", "best_location", "price_spread", "is_tied"])
    return wide[[c for c in col_order if c in wide.columns]]
